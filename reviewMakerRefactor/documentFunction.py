import re
from openpyxl import load_workbook

def find_version_number(sheet):
    # シート内の行数と列数を取得
    rows = sheet.max_row
    columns = sheet.max_column

    # 右下から左上方向にセルを読み込む
    for row in range(rows, 0, -1):
        for col in range(columns, 0, -1):
            cell = sheet.cell(row=row, column=col).value
            if cell:
                # セルの値が「vX.XX」または「X.XX」(Xは数字)の形式かどうかをチェック
                match = re.match(r'(v\d+\.\d+|\d+\.\d+)', str(cell))
                if match:
                    return match.group(0)
    return None

def get_version_from_excel(file_path):
    # Excelファイルを読み込む
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(f"エラー: ファイルを読み込めませんでした。({e})")
        return None

    # 「変更履歴」「更新履歴」シートがあるかどうかをチェック
    for sheet_name in ["変更履歴", "更新履歴"]:
        if sheet_name in wb:
            version_number = find_version_number(wb[sheet_name])
            if version_number:
                return version_number

    print("エラー: バージョン番号が見つかりませんでした。")
    return None

def load_file_content(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    except Exception as e:
        print(f"エラー: ファイルを読み込めませんでした。({e})")
        return None

def get_version_from_markdown(file_content):
    if file_content is None:
        return None

    # 正規表現で__vX.XX形式の文字列を検索
    pattern = re.compile(r'__v\d+\.\d+')
    matches = pattern.findall(file_content)

    if matches:
        # 最後のバージョン番号を返す
        return matches[-1]
    else:
        print("バージョン番号が見つかりませんでした。")
        return None

if __name__ == "__main__":
    file_path = "your_excel_file.xlsx"  # Excelファイルのパスを指定
    version_number = get_version_from_excel(file_path)
    if version_number:
        print(f"バージョン番号: {version_number}")
