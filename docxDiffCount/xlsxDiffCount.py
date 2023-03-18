from os import getcwd, path, remove
import openpyxl
import argparse

def cmpSheet(sheetOrg, sheetMod):
    if sheetOrg.max_row != sheetMod.max_row:
        return False

    if sheetOrg.max_column != sheetMod.max_column:
        return False

    maxRow = sheetOrg.max_row + 1
    maxColumn = sheetOrg.max_column + 1

    for column in range(1, maxRow):
        for row in range(1,maxColumn):
            if sheetOrg.cell(row, column).value != sheetMod.cell(row, column).value:
                print(f'different{column}:{row}_ {sheetOrg.cell(row, column)} x {sheetMod.cell(row, column)}')
                return False
            else:
                print(f'same {column}:{row}_ {sheetOrg.cell(row, column)} x {sheetMod.cell(row, column)}')
    return True

def cmpXlsxs(originalFileName, modifyFileName):
    ''' xlsxを比較する '''
    bookOrg = openpyxl.load_workbook(originalFileName)
    sheetsOrg = bookOrg.sheetnames

    bookMod = openpyxl.load_workbook(modifyFileName)
    sheetsMod = bookMod.sheetnames

    print(f'add {set(sheetsMod) - set(sheetsOrg)}')
    add = len(set(sheetsMod) - set(sheetsOrg))
    print(f'delete {set(sheetsOrg) - set(sheetsMod)}')
    delete = len(set(sheetsOrg) - set(sheetsMod))
    same = 0
    mod = 0
    for sheet in list(set(sheetsOrg) & set(sheetsMod)):
        if cmpSheet(bookOrg[sheet], bookMod[sheet]) == True:
            print(f'same: {sheet}')
            same = same +1
        else:
            print(f'mod: {sheet}')
            mod = mod + 1
    print(add, delete, mod, same)

if __name__ == '__main__':
    # get arg
    dir = getcwd() + '\\'
    argParser = argparse.ArgumentParser(description='docx diff conut')
    argParser.add_argument('--originalfile', default="D:\\pythontools\\pythontools\\pythontools\\countDocxDiff\\xlsxsample_org.xlsx", help='original file name')
    argParser.add_argument('--revisedfile', default="D:\\pythontools\\pythontools\\pythontools\\countDocxDiff\\xlsxsample_mod.xlsx", help='revised file name')

    args = argParser.parse_args()
    originalFileName = args.originalfile
    revisedFileName = args.revisedfile

    # create Compare docx
    compareFileName = cmpXlsxs(originalFileName, revisedFileName)
